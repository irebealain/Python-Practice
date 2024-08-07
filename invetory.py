import openpyxl

# Bringing the Excel sheet in the program using openpyxl modules

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Dictionary for the products per supplier
products_per_supplier = {}

# Dictionary for the value for products per supplier
total_value_per_supplier = {}

# Dictionary for the products with an inventory less than 50
products_under_50_inv = {}

for product_row in range(2, product_list.max_row + 1):

    # Getting supplier name
    supplier_name = product_list.cell(product_row, 4).value

    # Getting the inventory
    inventory = product_list.cell(product_row, 2).value

    # Getting the price for each product
    price = product_list.cell(product_row, 3).value

    # Getting the product number for each product
    product_number = product_list.cell(product_row, 1).value
    # Adding a new cell in each row to hold the value of the product
    inventory_price = product_list.cell(product_row, 5)
    # calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + price * inventory
    else:
        total_value_per_supplier[supplier_name] = price * inventory

    # Calculate all the products with an inventory under 50
    if inventory < 50:
        # Getting the inventories under 50 and add them into the dictionary of products with an inventory under 50
        products_under_50_inv[int(product_number)] = int(inventory)

    # Add the total for the inventory price
    inventory_price.value = price * inventory

# Saving a new inventory file with all the changes.
inv_file.save("inventory_with_total_inventory.xlsx")
# Printing all the dictionaries after saving a new file
print(products_under_50_inv)
print(total_value_per_supplier)
print(products_under_50_inv)